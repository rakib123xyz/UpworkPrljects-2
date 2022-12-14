#!/usr/bin/python3
# -*- coding: utf-8 -*-
import os
import sys
import logging

from selenium.webdriver.common.by import By
from esun_selenium import WebBrowser, send_discord_embed,notifyEmail

import config

import time
import openpyxl
import json

logger = logging.getLogger(__name__)
# ================== START PROGRAM IMPLEMENT ====================

def get_leads_excel():
        list_leads = []
        try:
            wb = openpyxl.load_workbook("data.xlsx")
            sheet = wb.active
            for i in range(1, sheet.max_row + 1):
                customer = sheet[f'A{i}'].value
                source = sheet[f'B{i}'].value
                created = sheet[f'C{i}'].value
                bucket = sheet[f'D{i}'].value

                obj = {'customer': customer, 'source': source, 'created': created, 'bucket': bucket}
                list_leads.append(obj)
        except:
            pass
        return list_leads

def update_leads_excel(list_leads):
        wb = openpyxl.Workbook()
        sheet = wb.active
        for i in range(0, len(list_leads)):
            sheet[f'A{i + 1}'] = list_leads[i].get("customer")
            sheet[f'B{i + 1}'] = list_leads[i].get("source")
            sheet[f'C{i + 1}'] = list_leads[i].get("created")
            sheet[f'D{i + 1}'] = list_leads[i].get("bucket")

        wb.save("data.xlsx")

def should_refresh(startTime,interval=30):
    started_at = startTime
    currentTime = time.time()
    interval = interval
    if (currentTime- started_at) > interval:
        return True
    else:
        return False


def get_data(browser,url):
    start_Time = time.time()
    global list_leads
    browser.get_url(url)
    time.sleep(1)
    print("Start login")
    input_username = browser.find_by_id("username")
    input_username.send_keys(username)

    btn_next = browser.find_by_id("signIn")
    time.sleep(1)
    browser.click_element(btn_next)

    time.sleep(1)
    input_password = browser.find_by_id("password")
    input_password.send_keys(password)

    btn_next = browser.find_by_id("signIn")
    time.sleep(1)
    browser.click_element(btn_next)
    inpt = input("Please type yes if login successful : ")
    time.sleep(5)
    
    print("Login successful")
    time.sleep(5)
    while True:
        try:

            if should_refresh(start_Time,interval=config.check_login_interval):
                print("Loging check..")

                #check logined? if not login. login again
                input_username = browser.find_by_id(locator="username",timeout=10)
                if input_username:

                    input_username.send_keys(username)

                    btn_next = browser.find_by_id("signIn")
                    time.sleep(1)
                    browser.click_element(btn_next)

                    time.sleep(1)
                    input_password = browser.find_by_id("password")
                    input_password.send_keys(password)

                    btn_next = browser.find_by_id("signIn")
                    time.sleep(1)
                    browser.click_element(btn_next)
                    start_Time = time.time()



            url_leads = "https://apps.vinmanager.com/CarDashboard/Pages/LeadManagement/ActiveLeads_WorkList.aspx"
            #url_leads = "file:///C:/Users/Rakib/Desktop/Untitled%20Page.html"
            browser.get_url(url_leads)
            time.sleep(1)
            # browser.switch_to_frame_by_id("cardashboardframe")
            #
            # time.sleep(.5)
            # browser.switch_to_frame_by_id("leftpaneframe")
            #
            # time.sleep(.5)

            #check data new lead

            #tbody_element = browser.find_all_by_xpath(locator='//div[@id="ctl00_ContentPlaceHolder1__LeadBucket"]//tbody')

            #
            # tbody = tbody_element[0]
            # print(tbody_element[0].text())

            trs = browser.find_elements_by_xpath(locator='//div[@id="ctl00_ContentPlaceHolder1__LeadBucket"]//tbody/tr')

            for tr in trs:

                try:

                    tds = tr.find_elements(By.TAG_NAME,"td")


                    # if "No leads to display" in tds[0].text:
                    #     print("No leads to display")
                    #     break
                    # if "This screen will auto-refresh every " in tr.text or "Customer Source" in tr.text:
                    #     continue
                    if len(tds) <= 1:
                        print("No leads to display")
                        break
                    if len(tds) > 1:

                        customer = tds[1].text
                        source = tds[2].text
                        created = tds[3].text
                        bucket = tds[4].text

                        obj = {'customer': customer, 'source': source, 'created': created, 'bucket': bucket}
                        if obj not in list_leads:
                            print("The web has new lead{}".format(obj))
                            list_leads.append(obj)
                            update_leads_excel(list_leads)

                            subject = "The web has new lead"
                            data = json.dumps(obj, indent=2)
                            mess = data
                            if type_notify == "G":
                                print("Send gmail")
                                notifyEmail(subject,mess,receiver_email)
                            else:
                                print("Send discord")
                                send_discord_embed(subject,
                                    mess,
                                    "",
                                    "","https://vinmanager.com",disc_hook)



                        else:
                            print(" already captured")
                except Exception as e1:
                    print(e1)


        except Exception as e:
            pass

        
        print("Time sleep to refresh")
        time.sleep(time_refresh)

def main():
    try:
        # Start browser
        browser = WebBrowser(timeout = 10,isMaximum = False, isDisableImage = False,isHeadless = False,)
        url = "https://apps.vinmanager.com/vinconnect"
        while True:        
            try:
                get_data(browser,url)
            except:
                pass

    except Exception as ex: # pylint: disable=W0703
        logger.info('An error has occurred.')
        logger.info('Contact author to resolve this issue')
        print(ex)
        import traceback
        print(traceback.format_exc())
        print('Pre ENTER to quit')

    print('------------------------------------ PROGRAM END ------------------------------------')
    print('\n\n\n\n')
    print('                     ===================== AUTHORS ======================')
    print('                     = Author:                             =')
    print('                     =                 =')
    print('                     =              =')
    print('                     = Thank you very much !                            =')
    print('                     ====================================================\n\n')
    print('Pre ENTER to quit')

    #sys.exit()


if __name__ == "__main__":
    global CURRENT_PATH, username, password, time_refresh, type_notify, user_gmail, pass_gmail, \
        receiver_email,disc_hook, list_leads
    # current folder path
    CURRENT_PATH = os.path.dirname(os.path.realpath(sys.argv[0]))

    username = config.username
    password = config.password
    time_refresh = config.time_refresh
    type_notify = config.type_notify
    user_gmail = config.user_gmail
    pass_gmail = config.pass_gmail
    receiver_email = config.receiver_email
    disc_hook = config.disc_hook
    #get log list old
    list_leads = get_leads_excel()

    main()    
